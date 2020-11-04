""" 
Get K-factor form spreadsheet and put into another spreadsheet
"""
version = "1.0.0"
# -----[ import stuff ]-----
import os
import openpyxl as xl



print(" -----[ K-Faktor festival ]-----")
print(" 1. MSA, Johan Sverdrup\n 2. MSB, Johan Sverdrup\n 3. MSA. Troll Blend\n 4. MSB,  Troll Blend") 

#inn = int(input())
invalid = True
orgfolders = ['G:\\R\\RAF_SM_FAG-DATA\\AUTOMASJ\\KALIBRERING\\3 - Trend\\Råolje målestasjon\\K-faktorer JS\\MSA JS.xlsx',
              'G:\\R\\RAF_SM_FAG-DATA\\AUTOMASJ\\KALIBRERING\\3 - Trend\\Råolje målestasjon\\K-faktorer JS\\MSB JS.xlsx',
              'G:\\R\\RAF_SM_FAG-DATA\\AUTOMASJ\\KALIBRERING\\3 - Trend\\Råolje målestasjon\\K-faktorer Trblend\\MSA TrBlend.xlsx',
              'G:\\R\\RAF_SM_FAG-DATA\\AUTOMASJ\\KALIBRERING\\3 - Trend\\Råolje målestasjon\\K-faktorer Trblend\\MSB TrBlend.xlsx']

innfolders = ['F:\\Scripts\\Kfaktorlogg\\MSA_1\\', 'F:\\Scripts\\Kfaktorlogg\\MSB_1\\', 'F:\\Scripts\\Kfaktorlogg\\MSA_14\\', 'F:\\Scripts\\Kfaktorlogg\\MSB_14\\']

#print(orgFile)
# Make list of excel files
runs = 0
print(len(orgfolders))
print(len(innfolders))
for orgFile in orgfolders:
    innfolder = innfolders[runs]
    print()
    print("-----[Working on folder]-----")
    print(orgFile)
    print()
    
    files = []

    mainwb = xl.load_workbook(orgFile)
    mainSheets = mainwb.sheetnames

    for f in os.listdir(innfolder):
        if f.endswith(".xlsx"):
            
            files.append(f)


    print(files)

    # -----[ find first merged cell ]-----

    #def findMerged (filename, fromCollumn):
    for file in files:
        merged = False
        print(" working on file: -----[ " + str(file) + " ]-----")
         
        filename = "file"
        wb = xl.load_workbook(innfolder + file)
        sheet = wb.worksheets[0]

        fromCollumn = 3
        while not merged:
            value = sheet.cell(row = fromCollumn, column = 3)
            
            
            celltype = type(value).__name__
            #print(celltype)
            #print(value)
            
            fromCollumn += 1

            #print(fromCollumn)
            if type(value).__name__ == 'MergedCell':
                fromCollumn = fromCollumn - 1
                break

        print("Merged cell: " + str(fromCollumn))
        # -----[ make dict ]-----
        count = 0
        values = {}
        for v in range (4, fromCollumn):
            rowValues = []
            for i in range(1, 16):
                
            
                cell = sheet.cell(row = v, column = i).value

                if cell == None:
                    continue
                else:
                    rowValues.append(cell)
            
            count += 1
            key = str(count)
                    
                
                
                #print(rowValues)
            values.update({key: rowValues})
            
            
        
        #print(values)

    # -----[ Find correct sheet ]-----
    
        #sheetIndex = file.split(" ")[1][0]
        
        temp = file.split(" ")[1]
        sheetIndex = temp.split("_")[0]
        print(sheetIndex)
        mainSheet = mainwb.worksheets[int(sheetIndex) - 1]

        print("Main sheet: " + str(mainSheet))

        # find empty column
        empty = False

        fromRow = 7
        while not empty:
            
            fromRow += 1
            value = mainSheet["A" + str(fromRow)].value
            #print(value)
            if value == None or value == "None":
                emptyCell = fromRow
                empty = True
                break 

        print("empty cell: " + str(emptyCell))

        # append dates
        dates = []
        for d in range(8, emptyCell):
            cell = mainSheet["A" + str(d)].value
            dates.append(cell)

        #print(dates)


        fromRowMain = emptyCell - 1
        
        for k in values:
            
            fromRowMain += 1
            data = values[k]
            
            y = 1
            
            
            if values[k][0] in dates:
                continue
            else:
                #print(data)
                for h in data:
                    
                    mainSheet.cell(row=int(fromRowMain), column = y).value = h
                    
                    y += 1

    splittFilnavn = orgFile.split("\\")
    nyttFilnavn = splittFilnavn[8]
    print(nyttFilnavn)                
    mainwb.save('F:\\Scripts\\Kfaktorlogg\\out\\' + str(nyttFilnavn))
    runs += 1
    print()
    print("Runs" + str(runs))
    print()        

    #mergedCell = findMerged("MSA_linje 3_1.xlsx", 6)


