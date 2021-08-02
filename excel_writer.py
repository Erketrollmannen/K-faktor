import os
from global_vars import folders
from datetime import date
try: 
    import openpyxl as xl
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ModuleNotFoundError:
    print("pip install --user openpyxl")
    os._exit(0)


def find_csvfiles(out_folder):
    csv_files = dict() # {folder: [files]}
    for key in folders:
        folder = f"{out_folder}/{key}"
        csv_files[key] = [file for file in os.listdir(folder) if file.endswith(".csv")]
    return csv_files

def read_csv_file(file):
    with open(file, "r") as f:
        lines = f.readlines()
    return lines

def find_wb_start(sheet):
    row = 1
    for row in range(1, 30):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value == None:
            continue
        elif "tidsmerke" in cell_value.lower():
            return row + 1
    raise Exception(f"Fant ikkje start på ark i : {sheet}")
        
def get_sheet_by_name(wb, line, station, oil_type):
    tmp_line = "0" if line == "10" else line
    if station == "MSA":
        for i in wb.worksheets:
            sheetname = i.title.split(",")
            meter = sheetname[1].split("-")[1]
            if f"Løp {line}" == sheetname[0] and tmp_line[0] == meter[1] and meter[0] in ("3", "4"):
                return i
    elif station == "MSB":
        for i in wb.worksheets:
            sheetname = i.title.split(",")
            meter = sheetname[1].split("-")[1]
            if f"Løp {line}" == sheetname[0] and tmp_line[0] == meter[1] and meter[0] in ("5", "6"):
                return i
    raise Exception(f"Fant ikkje passande worksheet, løp: {line}, på {station}, oljetype {oil_type}")


def data_to_excel(settings):
    workbooks = {"MSA_1": "MSA JS_template.xlsx", "MSB_1": "MSB JS_template.xlsx", "MSA_14": "MSA TrBlend_template.xlsx", "MSB_14": "MSB TrBlend_template.xlsx"}
    csv_files = find_csvfiles(settings["out_folder"])
    
    for key in  csv_files:
        key = key.split("/")[-1]
        wb = xl.load_workbook(workbooks[key])
        print(f"Working on {key}\n")
        for file in csv_files[key]:
            plt_data = list()
            lines = read_csv_file(f"{settings['out_folder']}/{key}/{file}") 
            tmp = file.split("_")
            try:
                station = tmp[0]
                line = tmp[1]
                oiltype = tmp[2].split(".")[0]
            except IndexError:
                continue

            sheet = get_sheet_by_name(wb, line, station, oiltype)
            start = find_wb_start(sheet)
            row = start
            for line in reversed(lines):
                line.strip()
                tmp = line.split(",")
                if len(tmp) != 11:
                    continue
                data = list()
                
                data.append(tmp[0])
                try:
                    data.append(int(tmp[1]))
                    data.append(float(tmp[2]))
                    data.append(float(tmp[3]))
                    data.append(float(tmp[4]))
                    data.append(float(tmp[5]))
                    data.append(float(tmp[6]))
                    data.append(float(tmp[7]))
                    data.append(float(tmp[8]))
                    data.append(float(tmp[9]))
                    data.append(int(tmp[10]))
                except ValueError:
                    continue
                for col in range(1, 12):
                    sheet.cell(row=row, column=col).value = data[col-1]

                plt_data.append([data[0], data[1]])
                data = list()
                row += 1

            print("Updating Table")
            try:
                table_ref = list(sheet.tables.values())[0]
            except IndexError:
                continue
            table_ref.ref = f"A{start-1}:K{row-1}"
            
        print(f"{settings['out_folder']}/{workbooks[key].split('_')[0]}_{date.today()}.xlsx")
        wb.save(f"{settings['out_folder']}/{workbooks[key].split('_')[0]}_{date.today()}.xlsx")
        
# if __name__ == "__main__":
#     data_to_excel()